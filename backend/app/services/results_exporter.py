"""Export simulation results as CSV or Excel."""
import csv
import io
from typing import Any


def export_csv(results: dict[str, Any]) -> str:
    """Export simulation results as CSV string."""
    output = io.StringIO()
    writer = csv.writer(output)

    # Stream Results section
    stream_results = results.get("stream_results", {})
    writer.writerow(["=== Stream Results ==="])
    writer.writerow(["Stream ID", "Temperature (°C)", "Pressure (kPa)", "Flow Rate (kg/s)", "Vapor Fraction", "Composition"])
    for stream_id, cond in stream_results.items():
        if not isinstance(cond, dict):
            continue
        comp = cond.get("composition", {})
        comp_str = "; ".join(f"{k}: {v:.4f}" for k, v in comp.items() if isinstance(v, (int, float))) if isinstance(comp, dict) else ""
        writer.writerow([
            stream_id,
            _fmt(cond.get("temperature")),
            _fmt(cond.get("pressure")),
            _fmt(cond.get("flowRate", cond.get("flow_rate"))),
            _fmt(cond.get("vapor_fraction")),
            comp_str,
        ])

    writer.writerow([])

    # Equipment Results section
    equipment_results = results.get("equipment_results", {})
    writer.writerow(["=== Equipment Results ==="])
    writer.writerow(["Equipment ID", "Type", "Duty (kW)", "Work (kW)", "Outlet Temp (°C)", "Outlet Pressure (kPa)"])
    for eq_id, eq_data in equipment_results.items():
        if not isinstance(eq_data, dict):
            continue
        writer.writerow([
            eq_id,
            eq_data.get("type", ""),
            _fmt(eq_data.get("duty")),
            _fmt(eq_data.get("work")),
            _fmt(eq_data.get("outlet_temperature")),
            _fmt(eq_data.get("outlet_pressure")),
        ])

    return output.getvalue()


def export_xlsx(results: dict[str, Any]) -> bytes:
    """Export simulation results as Excel workbook bytes."""
    from openpyxl import Workbook

    wb = Workbook()

    # Sheet 1: Streams
    ws_streams = wb.active
    ws_streams.title = "Streams"
    ws_streams.append(["Stream ID", "Temperature (°C)", "Pressure (kPa)", "Flow Rate (kg/s)", "Vapor Fraction", "Composition"])
    stream_results = results.get("stream_results", {})
    for stream_id, cond in stream_results.items():
        if not isinstance(cond, dict):
            continue
        comp = cond.get("composition", {})
        comp_str = "; ".join(f"{k}: {v:.4f}" for k, v in comp.items() if isinstance(v, (int, float))) if isinstance(comp, dict) else ""
        ws_streams.append([
            stream_id,
            cond.get("temperature"),
            cond.get("pressure"),
            cond.get("flowRate", cond.get("flow_rate")),
            cond.get("vapor_fraction"),
            comp_str,
        ])

    # Sheet 2: Equipment
    ws_equipment = wb.create_sheet("Equipment")
    ws_equipment.append(["Equipment ID", "Type", "Duty (kW)", "Work (kW)", "Outlet Temp (°C)", "Outlet Pressure (kPa)"])
    equipment_results = results.get("equipment_results", {})
    for eq_id, eq_data in equipment_results.items():
        if not isinstance(eq_data, dict):
            continue
        ws_equipment.append([
            eq_id,
            eq_data.get("type", ""),
            eq_data.get("duty"),
            eq_data.get("work"),
            eq_data.get("outlet_temperature"),
            eq_data.get("outlet_pressure"),
        ])

    # Sheet 3: Convergence
    ws_conv = wb.create_sheet("Convergence")
    conv_info = results.get("convergence_info", {})
    ws_conv.append(["Property", "Value"])
    ws_conv.append(["Iterations", conv_info.get("iterations", 0)])
    ws_conv.append(["Converged", str(conv_info.get("converged", False))])
    ws_conv.append(["Error", conv_info.get("error", 0)])
    ws_conv.append(["Mass Balance OK", str(conv_info.get("mass_balance_ok", "N/A"))])
    ws_conv.append(["Energy Balance OK", str(conv_info.get("energy_balance_ok", "N/A"))])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _fmt(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.4f}"
    return str(value)
