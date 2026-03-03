"""Parse uploaded CSV/XLSX/JSON files into insights-compatible format.

Supports:
- ProSim's own CSV/Excel export format (section markers)
- Generic CSV/Excel with heuristic column matching
- ProSim JSON (stream_results/equipment_results keys)
- DWSIM JSON (SimulationObjects keys)
- Unknown formats (raw context for AI)
"""

import csv
import io
import json
import logging
from typing import Any

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# ProSim CSV header → engine key mapping (P1 fix)
# ---------------------------------------------------------------------------
_PROSIM_HEADER_MAP: dict[str, str] = {
    "stream id": "_id",
    "temperature (°c)": "temperature",
    "temperature (c)": "temperature",
    "pressure (kpa)": "pressure",
    "flow rate (kg/s)": "mass_flow",
    "mass flow (kg/s)": "mass_flow",
    "vapor fraction": "vapor_fraction",
    "vapour fraction": "vapor_fraction",
    "composition": "composition",
    "enthalpy (kj/kg)": "enthalpy",
    "entropy (kj/kg/k)": "entropy",
    # Equipment headers
    "equipment id": "_id",
    "equipment name": "name",
    "equipment type": "type",
    "duty (kw)": "duty",
    "work (kw)": "work",
    "heat duty (kw)": "duty",
}

# ---------------------------------------------------------------------------
# Column keywords for heuristic matching (case-insensitive substring)
# Expanded with British spelling and HYSYS/PRO/II column names (P5 fix)
# ---------------------------------------------------------------------------
_STREAM_KEYWORDS = {
    "temperature", "temp", "t_c", "t(c)",
    "pressure", "press", "p_kpa", "p(kpa)", "p(psia)",
    "flow", "mass_flow", "massflow", "flowrate", "mass flow",
    "molar flow", "vol flow", "volumetric",
    "vapor_fraction", "vaporfraction", "vf",
    "vapour fraction", "vapour_fraction", "vapour",  # British spelling
    "enthalpy", "entropy",
    "composition", "phase",
    "mass density", "density", "molecular weight", "mw",
    "cp", "viscosity", "thermal conductivity", "surface tension",
    "heat capacity",
}
_EQUIPMENT_KEYWORDS = {
    "equipment", "duty", "work", "power", "efficiency", "area",
    "heat_duty", "q_kw", "w_kw", "delta_p", "pressure_drop",
    "head", "speed", "rpm", "stages", "reflux",
}
# Note: "type" intentionally NOT in _EQUIPMENT_KEYWORDS (D1/P4 fix)
# — stream tables often have a "type" column (vapor/liquid/two-phase)
_ALL_KEYWORDS = _STREAM_KEYWORDS | _EQUIPMENT_KEYWORDS

# Columns that strongly indicate equipment (not streams)
_STRONG_EQUIPMENT_MARKERS = {"duty", "work", "power", "head", "stages", "reflux", "speed", "rpm"}

# ---------------------------------------------------------------------------
# Unit system detection heuristics (D3 fix)
# ---------------------------------------------------------------------------

_UNIT_SYSTEMS = {
    "SI_ProSim": {"temperature": "°C", "pressure": "kPa", "mass_flow": "kg/s", "enthalpy": "kJ/kg"},
    "SI_DWSIM": {"temperature": "K", "pressure": "Pa", "mass_flow": "kg/s", "enthalpy": "J/mol"},
    "SI_HYSYS": {"temperature": "°C", "pressure": "kPa", "mass_flow": "kg/h", "enthalpy": "kJ/kg"},
    "Field": {"temperature": "°F", "pressure": "psia", "mass_flow": "lb/h", "enthalpy": "BTU/lb"},
}


def _detect_unit_system(stream_results: dict[str, Any]) -> tuple[str, list[str]]:
    """Detect likely unit system from value ranges. Returns (system_name, warnings)."""
    temps: list[float] = []
    pressures: list[float] = []
    for data in stream_results.values():
        if not isinstance(data, dict):
            continue
        # Fuzzy key lookup: match "temperature", "temperature (c)", "temp", etc.
        t = _fuzzy_get(data, ("temperature", "temp"))
        p = _fuzzy_get(data, ("pressure", "press"))
        if isinstance(t, (int, float)):
            temps.append(t)
        if isinstance(p, (int, float)):
            pressures.append(p)

    if not temps and not pressures:
        return "unknown", []

    warnings: list[str] = []
    system = "SI_ProSim"  # default

    if temps:
        avg_t = sum(temps) / len(temps)
        if avg_t > 200:
            # Likely Kelvin (DWSIM) or Fahrenheit
            if pressures:
                avg_p = sum(pressures) / len(pressures)
                if avg_p > 10000:
                    # Very high pressures → probably Pa (DWSIM SI)
                    system = "SI_DWSIM"
                    warnings.append(
                        "Detected DWSIM unit system (K, Pa). "
                        "Temperature and pressure values converted for display."
                    )
                elif avg_p < 500:
                    # Low pressures → probably psia (Field units)
                    system = "Field"
                    warnings.append(
                        "Detected field unit system (°F, psia). "
                        "Values shown as-is; verify unit consistency."
                    )
                else:
                    # Could be K with kPa — unusual but possible
                    system = "SI_DWSIM"
                    warnings.append(
                        "High temperature values detected — may be in Kelvin. "
                        "Verify unit system."
                    )
            else:
                warnings.append(
                    "High temperature values detected — may be in Kelvin or Fahrenheit. "
                    "Verify unit system."
                )

    return system, warnings


def _annotate_units_in_context(raw_context: str, unit_system: str) -> str:
    """Add unit system annotation to raw context for AI."""
    if unit_system == "unknown" or unit_system == "SI_ProSim":
        return raw_context
    units = _UNIT_SYSTEMS.get(unit_system, {})
    annotation = f"\n\n## Detected Unit System: {unit_system}\n"
    for prop, unit in units.items():
        annotation += f"- {prop}: {unit}\n"
    annotation += "NOTE: Interpret all numerical values using these units.\n"
    return raw_context + annotation


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def parse_insights_file(file_bytes: bytes, filename: str) -> dict[str, Any]:
    """Parse an uploaded file into insights-compatible format.

    Returns dict with keys: simulation_results, nodes, edges, raw_context,
    warnings, detected_unit_system, detected_property_package.
    """
    if not file_bytes:
        return {
            "simulation_results": {"stream_results": {}, "equipment_results": {}},
            "nodes": [],
            "edges": [],
            "raw_context": "",
            "warnings": ["Uploaded file is empty."],
            "detected_unit_system": "unknown",
            "detected_property_package": None,
        }

    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    warnings: list[str] = []

    # Encoding detection (P13 fix): try UTF-8-BOM, UTF-8, then Latin-1
    text_content: str | None = None
    if ext != "xlsx":
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                text_content = file_bytes.decode(enc)
                break
            except (UnicodeDecodeError, ValueError):
                continue
        if text_content is None:
            text_content = file_bytes.decode("utf-8", errors="replace")
            warnings.append("File encoding could not be detected; some characters may be corrupted.")

    if ext == "csv":
        result = _parse_csv(text_content, warnings)  # type: ignore[arg-type]
    elif ext == "xlsx":
        result = _parse_xlsx(file_bytes, warnings)
    elif ext == "xls":
        # P10 fix: .xls not supported by openpyxl
        warnings.append("Old .xls format is not supported. Please re-save as .xlsx in Excel.")
        raw = (text_content or file_bytes.decode("utf-8", errors="replace"))[:3000]
        result = {
            "simulation_results": {"stream_results": {}, "equipment_results": {}},
            "nodes": [],
            "edges": [],
            "raw_context": f"## Unsupported .xls file ({filename})\n{raw}",
            "warnings": warnings,
        }
    elif ext == "json":
        result = _parse_json(text_content, warnings)  # type: ignore[arg-type]
    else:
        warnings.append(f"Unsupported file extension '.{ext}'. AI will analyze raw text.")
        raw = (text_content or file_bytes.decode("utf-8", errors="replace"))[:5000]
        result = {
            "simulation_results": {"stream_results": {}, "equipment_results": {}},
            "nodes": [],
            "edges": [],
            "raw_context": f"## Raw File Content ({filename})\n{raw}",
            "warnings": warnings,
        }

    # Unit system detection (D3 fix)
    sr = result.get("simulation_results", {}).get("stream_results", {})
    unit_system, unit_warnings = _detect_unit_system(sr)
    result["warnings"] = result.get("warnings", []) + unit_warnings
    result["detected_unit_system"] = unit_system

    # Annotate raw_context with unit info
    if result.get("raw_context"):
        result["raw_context"] = _annotate_units_in_context(result["raw_context"], unit_system)

    # Property package detection from JSON data
    result.setdefault("detected_property_package", None)

    return result


# ---------------------------------------------------------------------------
# CSV parser
# ---------------------------------------------------------------------------

def _parse_csv(text: str, warnings: list[str]) -> dict[str, Any]:
    # Check for ProSim native format (section markers from results_exporter.py)
    if "=== Stream Results ===" in text or "=== Equipment Results ===" in text:
        return _parse_prosim_csv(text, warnings)

    # Heuristic column matching
    return _parse_generic_csv(text, warnings)


def _parse_prosim_csv(text: str, warnings: list[str]) -> dict[str, Any]:
    """Parse ProSim's own CSV export with section markers."""
    stream_results: dict[str, Any] = {}
    equipment_results: dict[str, Any] = {}

    lines = text.strip().split("\n")
    section = None
    headers: list[str] = []
    mapped_headers: list[str] = []

    for line in lines:
        stripped = line.strip()
        if "=== Stream Results ===" in stripped:
            section = "streams"
            headers = []
            mapped_headers = []
            continue
        elif "=== Equipment Results ===" in stripped:
            section = "equipment"
            headers = []
            mapped_headers = []
            continue
        elif stripped.startswith("==="):
            section = None
            continue

        if not stripped:
            continue

        row = list(csv.reader([stripped]))[0]

        if section in ("streams", "equipment"):
            if not headers:
                headers = [h.strip().lower() for h in row]
                # P1 fix: map ProSim headers to engine key names
                mapped_headers = [_PROSIM_HEADER_MAP.get(h, h) for h in headers]
                continue
            if len(row) >= 2:
                rid = row[0].strip()
                data: dict[str, Any] = {}
                for i, h in enumerate(mapped_headers):
                    if h == "_id":
                        continue  # skip redundant ID column
                    if i < len(row) and row[i].strip():
                        data[h] = _try_number(row[i].strip())
                if section == "streams":
                    stream_results[rid] = data
                else:
                    equipment_results[rid] = data

    nodes = _build_nodes_from_equipment(equipment_results)
    # P16 fix: brief metadata only (structured data already in simulation_results)
    raw_context = (
        f"## ProSim CSV Export\n"
        f"Streams: {len(stream_results)}, Equipment: {len(equipment_results)}\n"
        f"Format: ProSim Cloud native export"
    )

    return {
        "simulation_results": {
            "stream_results": stream_results,
            "equipment_results": equipment_results,
        },
        "nodes": nodes,
        "edges": [],
        "raw_context": raw_context,
        "warnings": warnings,
    }


def _parse_generic_csv(text: str, warnings: list[str]) -> dict[str, Any]:
    """Parse generic CSV using heuristic column matching."""
    lines = text.strip().split("\n")
    reader = list(csv.reader(lines))

    # Find header row: first row with 3+ recognized keywords
    header_idx = -1
    headers: list[str] = []
    for i, row in enumerate(reader):
        lower_row = [c.strip().lower() for c in row]
        matches = sum(1 for c in lower_row if any(kw in c for kw in _ALL_KEYWORDS))
        if matches >= 3:
            header_idx = i
            headers = lower_row
            break

    if header_idx < 0:
        warnings.append("Could not auto-detect column headers. AI will analyze raw data.")
        raw = "\n".join(lines[:50])
        return {
            "simulation_results": {"stream_results": {}, "equipment_results": {}},
            "nodes": [],
            "edges": [],
            "raw_context": f"## CSV Data (headers not detected)\n{raw}",
            "warnings": warnings,
        }

    # Parse data rows
    stream_results: dict[str, Any] = {}
    equipment_results: dict[str, Any] = {}
    raw_context_lines = [",".join(reader[header_idx])]

    for row in reader[header_idx + 1:]:
        if not row or all(c.strip() == "" for c in row):
            continue
        entry: dict[str, Any] = {}
        for j, h in enumerate(headers):
            if j < len(row) and row[j].strip():
                entry[h] = _try_number(row[j].strip())

        raw_context_lines.append(",".join(row))
        rid = row[0].strip() if row else f"row-{len(stream_results) + len(equipment_results)}"

        # D1/P4 fix: classify using strong equipment markers only
        # (not "type" which is ambiguous — stream tables have "type" columns)
        has_strong_eq = any(
            any(kw in h for kw in _STRONG_EQUIPMENT_MARKERS)
            for h in entry
        )
        if has_strong_eq:
            equipment_results[rid] = entry
        else:
            stream_results[rid] = entry

    nodes = _build_nodes_from_equipment(equipment_results)
    raw_context = f"## CSV Data ({len(stream_results)} streams, {len(equipment_results)} equipment)\n"
    raw_context += "### Headers + Sample Rows\n" + "\n".join(raw_context_lines[:50])

    return {
        "simulation_results": {
            "stream_results": stream_results,
            "equipment_results": equipment_results,
        },
        "nodes": nodes,
        "edges": [],
        "raw_context": raw_context,
        "warnings": warnings,
    }


# ---------------------------------------------------------------------------
# Excel parser
# ---------------------------------------------------------------------------

def _parse_xlsx(file_bytes: bytes, warnings: list[str]) -> dict[str, Any]:
    """Parse Excel file using openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        warnings.append("openpyxl not installed — cannot parse Excel files.")
        return {
            "simulation_results": {"stream_results": {}, "equipment_results": {}},
            "nodes": [],
            "edges": [],
            "raw_context": "Excel parsing requires openpyxl.",
            "warnings": warnings,
        }

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    stream_results: dict[str, Any] = {}
    equipment_results: dict[str, Any] = {}
    raw_context_parts: list[str] = ["## Excel Workbook"]
    extra_sheets: list[str] = []

    # P7 fix: expanded with common simulator sheet names + substring matching
    stream_substrings = {"stream", "material", "mass balance", "hmb", "heat and material"}
    equip_substrings = {"equipment", "unit op", "result", "summary"}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        lower_name = sheet_name.strip().lower()

        # Substring matching instead of exact set membership
        is_stream = any(sub in lower_name for sub in stream_substrings)
        is_equip = not is_stream and any(sub in lower_name for sub in equip_substrings)

        if is_stream:
            parsed = _parse_sheet_rows(rows, warnings)
            stream_results.update(parsed)
            raw_context_parts.append(f"### Sheet '{sheet_name}': {len(parsed)} streams")
        elif is_equip:
            parsed = _parse_sheet_rows(rows, warnings)
            equipment_results.update(parsed)
            raw_context_parts.append(f"### Sheet '{sheet_name}': {len(parsed)} equipment")
        else:
            # Unknown sheet — include headers in raw_context
            headers_row = rows[0] if rows else []
            header_str = ", ".join(str(h) for h in headers_row if h is not None)
            extra_sheets.append(f"Sheet '{sheet_name}' ({len(rows)} rows): {header_str}")
            # Try heuristic parse anyway
            parsed = _parse_sheet_rows(rows, warnings)
            if parsed:
                # D1 fix: use strong equipment markers for classification
                has_eq = any(
                    isinstance(v, dict) and any(
                        m in v for m in _STRONG_EQUIPMENT_MARKERS
                    )
                    for v in parsed.values()
                )
                if has_eq:
                    equipment_results.update(parsed)
                else:
                    stream_results.update(parsed)

    wb.close()

    if extra_sheets:
        raw_context_parts.append("### Other Sheets\n" + "\n".join(extra_sheets))

    nodes = _build_nodes_from_equipment(equipment_results)
    raw_context_parts.insert(1, f"Streams: {len(stream_results)}, Equipment: {len(equipment_results)}")

    return {
        "simulation_results": {
            "stream_results": stream_results,
            "equipment_results": equipment_results,
        },
        "nodes": nodes,
        "edges": [],
        "raw_context": "\n".join(raw_context_parts),
        "warnings": warnings,
    }


def _parse_sheet_rows(rows: list[tuple], warnings: list[str]) -> dict[str, Any]:
    """Parse Excel sheet rows using heuristic header detection."""
    if not rows:
        return {}

    # Find header row
    header_idx = -1
    headers: list[str] = []
    for i, row in enumerate(rows):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        matches = sum(1 for c in cells if any(kw in c for kw in _ALL_KEYWORDS))
        if matches >= 2:
            header_idx = i
            headers = cells
            break

    if header_idx < 0:
        # Use first row as headers if it has text
        first = rows[0]
        if first and any(c is not None and not _is_number(str(c)) for c in first):
            header_idx = 0
            headers = [str(c).strip().lower() if c is not None else f"col{j}" for j, c in enumerate(first)]
        else:
            return {}

    result: dict[str, Any] = {}
    for row in rows[header_idx + 1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        rid = str(row[0]).strip() if row[0] is not None else f"row-{len(result)}"
        entry: dict[str, Any] = {}
        for j, h in enumerate(headers):
            if j < len(row) and row[j] is not None:
                val = row[j]
                if isinstance(val, (int, float)):
                    entry[h] = val
                else:
                    entry[h] = _try_number(str(val).strip())
        result[rid] = entry

    return result


# ---------------------------------------------------------------------------
# JSON parser
# ---------------------------------------------------------------------------

def _parse_json(text: str, warnings: list[str]) -> dict[str, Any]:
    """Parse JSON file — detect ProSim, DWSIM, or generic format."""
    try:
        data = json.loads(text)
    except json.JSONDecodeError as e:
        warnings.append(f"Invalid JSON: {e}")
        return {
            "simulation_results": {"stream_results": {}, "equipment_results": {}},
            "nodes": [],
            "edges": [],
            "raw_context": f"## Invalid JSON\n{text[:5000]}",
            "warnings": warnings,
        }

    if not isinstance(data, dict):
        warnings.append("JSON root is not an object.")
        return {
            "simulation_results": {"stream_results": {}, "equipment_results": {}},
            "nodes": [],
            "edges": [],
            "raw_context": f"## JSON Data\n{json.dumps(data, default=str)[:5000]}",
            "warnings": warnings,
        }

    # ProSim native format
    if "stream_results" in data or "equipment_results" in data:
        sr = data.get("stream_results", {})
        er = data.get("equipment_results", {})
        nodes = _build_nodes_from_equipment(er)
        return {
            "simulation_results": {"stream_results": sr, "equipment_results": er},
            "nodes": data.get("nodes", nodes),
            "edges": data.get("edges", []),
            "raw_context": f"## ProSim JSON\nStreams: {len(sr)}, Equipment: {len(er)}",
            "warnings": warnings,
        }

    # Check for nested results key (from simulation response)
    if "results" in data and isinstance(data["results"], dict):
        inner = data["results"]
        if "stream_results" in inner or "equipment_results" in inner:
            sr = inner.get("stream_results", {})
            er = inner.get("equipment_results", {})
            nodes = _build_nodes_from_equipment(er)
            return {
                "simulation_results": {"stream_results": sr, "equipment_results": er},
                "nodes": nodes,
                "edges": [],
                "raw_context": f"## ProSim JSON (nested results)\nStreams: {len(sr)}, Equipment: {len(er)}",
                "warnings": warnings,
            }

    # DWSIM format
    if "SimulationObjects" in data:
        return _parse_dwsim_json(data, warnings)

    # Generic JSON — dump key/value pairs as context
    warnings.append("Could not auto-detect JSON format. AI will analyze raw data.")
    raw_parts: list[str] = ["## JSON Data (unknown format)"]
    count = 0
    for k, v in data.items():
        if count >= 100:
            raw_parts.append("... (truncated)")
            break
        if isinstance(v, (dict, list)):
            raw_parts.append(f"- {k}: {json.dumps(v, default=str)[:200]}")
        else:
            raw_parts.append(f"- {k}: {v}")
        count += 1

    return {
        "simulation_results": {"stream_results": {}, "equipment_results": {}},
        "nodes": [],
        "edges": [],
        "raw_context": "\n".join(raw_parts),
        "warnings": warnings,
    }


def _parse_dwsim_json(data: dict, warnings: list[str]) -> dict[str, Any]:
    """Parse DWSIM JSON export format."""
    stream_results: dict[str, Any] = {}
    equipment_results: dict[str, Any] = {}
    nodes: list[dict[str, Any]] = []
    detected_pp: str | None = None

    # Detect property package from DWSIM options
    options = data.get("Options", data.get("FlowsheetOptions", {}))
    if isinstance(options, dict):
        pp = options.get("PropertyPackage", options.get("SelectedPropertyPackage"))
        if pp:
            detected_pp = str(pp)

    sim_objs = data.get("SimulationObjects", {})
    if not isinstance(sim_objs, dict):
        sim_objs = {}

    # P11 fix: iterate once, classify by ObjectType
    _STREAM_TYPES = {"MaterialStream", "material stream", "Material Stream"}
    _SKIP_TYPES = {"EnergyStream", "energy stream", "Energy Stream"}

    for oid, odata in sim_objs.items():
        if not isinstance(odata, dict):
            continue
        otype = odata.get("ObjectType", odata.get("Type", ""))

        if otype in _STREAM_TYPES:
            # Extract stream properties
            entry: dict[str, Any] = {}
            props = odata.get("Properties", odata)
            if isinstance(props, dict):
                entry["temperature"] = props.get("Temperature", props.get("temperature"))
                entry["pressure"] = props.get("Pressure", props.get("pressure"))
                entry["mass_flow"] = props.get("MassFlow", props.get("massflow", props.get("mass_flow")))
                entry["vapor_fraction"] = props.get("VaporFraction", props.get("vapor_fraction"))
            entry = {k: v for k, v in entry.items() if v is not None}
            if entry:
                stream_results[oid] = entry

        elif otype in _SKIP_TYPES:
            continue

        elif otype:
            # Equipment object
            eq_entry: dict[str, Any] = {"type": otype}
            props = odata.get("Properties", odata)
            if isinstance(props, dict):
                for k, v in props.items():
                    if isinstance(v, (int, float)):
                        eq_entry[k] = v
            equipment_results[oid] = eq_entry
            nodes.append({
                "id": oid,
                "type": otype,
                "name": odata.get("Name", oid),
                "parameters": {},
            })

    warnings.append(
        "DWSIM properties are in SI units (K, Pa, kg/s). "
        "Temperature and pressure values may differ from ProSim conventions."
    )

    raw_context = (
        f"## DWSIM JSON Export\n"
        f"Streams: {len(stream_results)}, Equipment: {len(equipment_results)}\n"
        f"NOTE: DWSIM uses SI units — temperatures in Kelvin, pressures in Pascals."
    )

    result: dict[str, Any] = {
        "simulation_results": {
            "stream_results": stream_results,
            "equipment_results": equipment_results,
        },
        "nodes": nodes,
        "edges": [],
        "raw_context": raw_context,
        "warnings": warnings,
    }
    if detected_pp:
        result["detected_property_package"] = detected_pp
    return result


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _fuzzy_get(data: dict[str, Any], prefixes: tuple[str, ...]) -> Any:
    """Get a value from a dict by matching key prefixes (handles 'temperature (c)' etc.)."""
    for key, val in data.items():
        for pfx in prefixes:
            if key == pfx or key.startswith(pfx + " ") or key.startswith(pfx + "("):
                return val
    return None


def _build_nodes_from_equipment(equipment_results: dict[str, Any]) -> list[dict[str, Any]]:
    """Build synthetic node list from equipment results for AI context."""
    nodes: list[dict[str, Any]] = []
    for eid, data in equipment_results.items():
        if not isinstance(data, dict):
            continue
        etype = data.get("type", data.get("equipment", data.get("equipment_type", "Unknown")))
        name = data.get("name", data.get("equipment_name", eid))
        nodes.append({
            "id": eid,
            "type": str(etype),
            "name": str(name),
            "parameters": {},
        })
    return nodes


def _try_number(s: str) -> int | float | str:
    """Try to convert string to number, return original if not possible."""
    try:
        if "." in s or ("e" in s.lower() and any(c.isdigit() for c in s)):
            return float(s)
        return int(s)
    except (ValueError, TypeError):
        return s


def _is_number(s: str) -> bool:
    try:
        float(s)
        return True
    except (ValueError, TypeError):
        return False
